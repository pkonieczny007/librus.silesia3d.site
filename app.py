from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import os
from datetime import datetime
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import zipfile
import shutil

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'output'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Tworzenie folderów jeśli nie istnieją
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
os.makedirs('static', exist_ok=True)

HANDLOWCY_FILE = 'handlowcy.xlsx'

# Rejestracja czcionki obsługującej polskie znaki
try:
    pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
    DEFAULT_FONT = 'DejaVuSans'
    DEFAULT_FONT_BOLD = 'DejaVuSans-Bold'
except:
    try:
        pdfmetrics.registerFont(TTFont('DejaVuSans', 'C:\\Windows\\Fonts\\DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', 'C:\\Windows\\Fonts\\DejaVuSans-Bold.ttf'))
        DEFAULT_FONT = 'DejaVuSans'
        DEFAULT_FONT_BOLD = 'DejaVuSans-Bold'
    except:
        DEFAULT_FONT = 'Helvetica'
        DEFAULT_FONT_BOLD = 'Helvetica-Bold'

# Globalna zmienna do przechowywania danych sesji
session_data = {}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nie wybrano pliku'}), 400
    
    if file and file.filename.endswith('.csv'):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        return jsonify({'success': True, 'filename': filename})
    
    return jsonify({'error': 'Nieprawidłowy format pliku'}), 400

@app.route('/process', methods=['POST'])
def process_data():
    data = request.json
    filename = data.get('filename')
    date_from = data.get('date_from')
    date_to = data.get('date_to')
    time_from = data.get('time_from', '00:00')
    time_to = data.get('time_to', '23:59')
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    try:
        # Wczytanie pliku CSV
        df = pd.read_csv(filepath, delimiter=';', encoding='utf-8')
        
        # Konwersja daty płatności
        df['data płatności'] = pd.to_datetime(df['data płatności'], format='%d.%m.%Y %H:%M')
        
        # Konwersja kwoty (zamiana przecinka na kropkę)
        df['kwota'] = df['kwota'].astype(str).str.replace(',', '.').astype(float)
        
        # Tworzenie zakresów dat
        datetime_from = datetime.strptime(f"{date_from} {time_from}", "%Y-%m-%d %H:%M")
        datetime_to = datetime.strptime(f"{date_to} {time_to}", "%Y-%m-%d %H:%M")
        
        # Filtrowanie danych
        df_filtered = df[(df['data płatności'] >= datetime_from) & 
                        (df['data płatności'] <= datetime_to)].copy()
        
        # Wczytanie danych handlowców
        df_handlowcy = pd.read_excel(HANDLOWCY_FILE)
        
        # Tworzenie pliku Excel
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_excel = f"raport_{timestamp}.xlsx"
        output_path = os.path.join('static', output_excel)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Zakładka 1: Cały plik CSV
            df.to_excel(writer, sheet_name='Cały plik', index=False)
            
            # Zakładka 2: Dane do wyliczeń
            df_filtered.to_excel(writer, sheet_name='DANE DO WYLICZEŃ', index=False)
            
            # Zakładka 3: Podsumowanie grup
            df_grouped = df_filtered.groupby('Grupa zajęciowa')['kwota'].sum().reset_index()
            df_grouped.columns = ['Grupa zajęciowa', 'Suma kwot']
            df_grouped['10% z sumy'] = df_grouped['Suma kwot'] * 0.1
            
            # Przypisanie handlowców
            df_grouped = df_grouped.merge(df_handlowcy, on='Grupa zajęciowa', how='left')
            df_grouped['Handlowiec'] = df_grouped['Handlowiec'].fillna('BRAK PRZYPISANIA')
            
            df_grouped[['Grupa zajęciowa', 'Suma kwot', '10% z sumy', 'Handlowiec']].to_excel(
                writer, sheet_name='PODSUMOWANIE GRUP', index=False
            )
            
            # Zakładka 4: Podsumowanie handlowców
            df_handlowcy_sum = df_grouped.groupby('Handlowiec')['10% z sumy'].sum().reset_index()
            df_handlowcy_sum.columns = ['Handlowiec', 'Suma prowizji (10%)']
            df_handlowcy_sum.to_excel(writer, sheet_name='PODSUMOWANIE HANDLOWCÓW', index=False)
        
        # Formatowanie Excela
        format_excel(output_path, df_grouped, df_handlowcy_sum)
        
        # Tworzenie PDF
        output_pdf = f"raport_{timestamp}.pdf"
        pdf_path = os.path.join('static', output_pdf)
        create_pdf(pdf_path, df_grouped, df_handlowcy_sum, date_from, date_to, time_from, time_to)
        
        # Zapisanie danych do sesji dla późniejszego użycia
        session_id = timestamp
        session_data[session_id] = {
            'df_filtered': df_filtered,
            'df_grouped': df_grouped,
            'df_handlowcy_sum': df_handlowcy_sum,
            'date_from': date_from,
            'date_to': date_to,
            'time_from': time_from,
            'time_to': time_to
        }
        
        # Lista handlowców do wyboru
        handlowcy_list = df_handlowcy_sum['Handlowiec'].tolist()
        
        return jsonify({
            'success': True,
            'excel_file': output_excel,
            'pdf_file': output_pdf,
            'date_range': f"{date_from} {time_from} - {date_to} {time_to}",
            'session_id': session_id,
            'handlowcy': handlowcy_list
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate_handlowiec_reports', methods=['POST'])
def generate_handlowiec_reports():
    """Generowanie raportów dla wybranych handlowców"""
    data = request.json
    session_id = data.get('session_id')
    selected_handlowcy = data.get('selected_handlowcy', [])
    
    if not session_id or session_id not in session_data:
        return jsonify({'error': 'Brak danych sesji. Wygeneruj najpierw główny raport.'}), 400
    
    if not selected_handlowcy:
        return jsonify({'error': 'Nie wybrano żadnego handlowca.'}), 400
    
    try:
        sess = session_data[session_id]
        df_filtered = sess['df_filtered']
        df_grouped = sess['df_grouped']
        date_from = sess['date_from']
        date_to = sess['date_to']
        time_from = sess['time_from']
        time_to = sess['time_to']
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Folder na pliki handlowców
        handlowcy_folder = os.path.join(app.config['OUTPUT_FOLDER'], f'handlowcy_{timestamp}')
        os.makedirs(handlowcy_folder, exist_ok=True)
        
        generated_files = []
        
        for handlowiec in selected_handlowcy:
            # Filtrowanie danych dla handlowca
            df_handlowiec = df_grouped[df_grouped['Handlowiec'] == handlowiec].copy()
            
            if df_handlowiec.empty:
                continue
            
            # Bezpieczna nazwa pliku
            safe_name = secure_filename(handlowiec.replace(' ', '_'))
            
            # Tworzenie Excel dla handlowca
            excel_filename = f"raport_{safe_name}_{timestamp}.xlsx"
            excel_path = os.path.join(handlowcy_folder, excel_filename)
            create_handlowiec_excel(excel_path, df_handlowiec, df_filtered, handlowiec, 
                                   date_from, date_to, time_from, time_to)
            
            # Tworzenie PDF dla handlowca
            pdf_filename = f"raport_{safe_name}_{timestamp}.pdf"
            pdf_path = os.path.join(handlowcy_folder, pdf_filename)
            create_handlowiec_pdf(pdf_path, df_handlowiec, df_filtered, handlowiec,
                                 date_from, date_to, time_from, time_to)
            
            generated_files.append({
                'handlowiec': handlowiec,
                'excel': excel_filename,
                'pdf': pdf_filename
            })
        
        # Tworzenie archiwum ZIP
        zip_filename = f"raporty_handlowcow_{timestamp}.zip"
        zip_path = os.path.join('static', zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(handlowcy_folder):
                for file in files:
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, file)
        
        # Kopiowanie plików do static dla pojedynczego pobierania
        for file in generated_files:
            src_excel = os.path.join(handlowcy_folder, file['excel'])
            src_pdf = os.path.join(handlowcy_folder, file['pdf'])
            shutil.copy(src_excel, os.path.join('static', file['excel']))
            shutil.copy(src_pdf, os.path.join('static', file['pdf']))
        
        # Czyszczenie folderu tymczasowego
        shutil.rmtree(handlowcy_folder)
        
        return jsonify({
            'success': True,
            'zip_file': zip_filename,
            'files': generated_files
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/upload_handlowiec_excel', methods=['POST'])
def upload_handlowiec_excel():
    """Wgrywanie poprawionego pliku Excel handlowca i generowanie nowego PDF"""
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku'}), 400
    
    file = request.files['file']
    handlowiec_name = request.form.get('handlowiec_name', 'Handlowiec')
    
    if file.filename == '':
        return jsonify({'error': 'Nie wybrano pliku'}), 400
    
    if not file.filename.endswith('.xlsx'):
        return jsonify({'error': 'Wymagany format pliku: .xlsx'}), 400
    
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Zapisanie przesłanego pliku
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Wczytanie danych z pliku Excel
        df_podsumowanie = pd.read_excel(filepath, sheet_name='PODSUMOWANIE')
        df_szczegoly = pd.read_excel(filepath, sheet_name='SZCZEGÓŁY')
        
        # Generowanie nowego PDF
        safe_name = secure_filename(handlowiec_name.replace(' ', '_'))
        pdf_filename = f"raport_{safe_name}_poprawiony_{timestamp}.pdf"
        pdf_path = os.path.join('static', pdf_filename)
        
        create_handlowiec_pdf_from_excel(pdf_path, df_podsumowanie, df_szczegoly, handlowiec_name)
        
        return jsonify({
            'success': True,
            'pdf_file': pdf_filename,
            'message': f'PDF dla {handlowiec_name} został wygenerowany pomyślnie.'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def create_handlowiec_excel(filepath, df_handlowiec, df_filtered, handlowiec_name, 
                           date_from, date_to, time_from, time_to):
    """Tworzenie pliku Excel dla konkretnego handlowca"""
    
    # Pobranie grup przypisanych do handlowca
    grupy_handlowca = df_handlowiec['Grupa zajęciowa'].tolist()
    
    # Filtrowanie szczegółowych danych dla grup handlowca
    df_szczegoly = df_filtered[df_filtered['Grupa zajęciowa'].isin(grupy_handlowca)].copy()
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Zakładka 1: Podsumowanie
        df_podsumowanie = df_handlowiec[['Grupa zajęciowa', 'Suma kwot', '10% z sumy']].copy()
        df_podsumowanie.to_excel(writer, sheet_name='PODSUMOWANIE', index=False)
        
        # Zakładka 2: Szczegóły (wszystkie wpłaty dla grup handlowca)
        df_szczegoly.to_excel(writer, sheet_name='SZCZEGÓŁY', index=False)
    
    # Formatowanie pliku
    format_handlowiec_excel(filepath, df_handlowiec, handlowiec_name, date_from, date_to, time_from, time_to)

def format_handlowiec_excel(filepath, df_handlowiec, handlowiec_name, date_from, date_to, time_from, time_to):
    """Formatowanie pliku Excel handlowca"""
    wb = openpyxl.load_workbook(filepath)
    
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sum_font = Font(bold=True, size=11, color="1F4788")
    sum_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # BEZ wstawiania dodatkowych wierszy nagłówkowych!
        # Formatowanie nagłówków tabeli (teraz w wierszu 1)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Formatowanie komórek danych
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border
                if isinstance(cell.value, (int, float)) and cell.column in [2, 3]:
                    cell.number_format = '#,##0.00 "PLN"'
        
        # Automatyczne dopasowanie szerokości kolumn
        for column_cells in ws.columns:
            length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        if len(cell_value) > length:
                            length = len(cell_value)
                except:
                    pass
            adjusted_width = min(length + 4, 80)
            ws.column_dimensions[column].width = adjusted_width
        
        ws.freeze_panes = 'A2'
    
    # Dodanie sumy w zakładce PODSUMOWANIE
    ws_pods = wb['PODSUMOWANIE']
    last_row = ws_pods.max_row + 1
    
    suma_kwot = df_handlowiec['Suma kwot'].sum()
    suma_prowizji = df_handlowiec['10% z sumy'].sum()
    
    ws_pods[f'A{last_row}'] = 'SUMA CAŁKOWITA:'
    ws_pods[f'B{last_row}'] = suma_kwot
    ws_pods[f'C{last_row}'] = suma_prowizji
    
    for col in ['A', 'B', 'C']:
        cell = ws_pods[f'{col}{last_row}']
        cell.font = sum_font
        cell.fill = sum_fill
        cell.border = border
        if col in ['B', 'C']:
            cell.number_format = '#,##0.00 "PLN"'
    
    wb.save(filepath)


def create_handlowiec_pdf(filepath, df_handlowiec, df_filtered, handlowiec_name,
                         date_from, date_to, time_from, time_to):
    """Tworzenie raportu PDF dla konkretnego handlowca"""
    
    # Pobranie grup przypisanych do handlowca
    grupy_handlowca = df_handlowiec['Grupa zajęciowa'].tolist()
    
    # Filtrowanie szczegółowych danych
    df_szczegoly = df_filtered[df_filtered['Grupa zajęciowa'].isin(grupy_handlowca)].copy()
    
    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(A4),
        topMargin=15*mm,
        bottomMargin=20*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Logo
    logo_path = os.path.join('static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=40*mm, height=40*mm)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 5))
        except:
            pass
    
    # Style
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontName=DEFAULT_FONT_BOLD,
        fontSize=18, alignment=TA_CENTER, spaceAfter=12, textColor=colors.HexColor('#1f4788')
    )
    subtitle_style = ParagraphStyle(
        'SubTitle', parent=styles['Heading2'], fontName=DEFAULT_FONT_BOLD,
        fontSize=14, alignment=TA_CENTER, spaceAfter=8, textColor=colors.HexColor('#333333')
    )
    date_style = ParagraphStyle(
        'DateStyle', parent=styles['Normal'], fontName=DEFAULT_FONT,
        fontSize=11, alignment=TA_CENTER, spaceAfter=20
    )
    section_style = ParagraphStyle(
        'SectionHeader', parent=styles['Heading2'], fontName=DEFAULT_FONT_BOLD,
        fontSize=14, textColor=colors.HexColor('#1f4788'), spaceAfter=10
    )
    cell_style = ParagraphStyle(
        'CellStyle', parent=styles['Normal'], fontName=DEFAULT_FONT,
        fontSize=9, leading=11
    )
    cell_style_bold = ParagraphStyle(
        'CellStyleBold', parent=styles['Normal'], fontName=DEFAULT_FONT_BOLD,
        fontSize=10, textColor=colors.HexColor('#1f4788'), leading=12
    )
    
    # Tytuł
    elements.append(Paragraph('RAPORT WPŁAT', title_style))
    elements.append(Paragraph(f'Handlowiec: {handlowiec_name}', subtitle_style))
    elements.append(Spacer(1, 10))
    
    date_range = f"Okres: {date_from} {time_from} - {date_to} {time_to}"
    elements.append(Paragraph(date_range, date_style))
    elements.append(Spacer(1, 15))
    
    # === SEKCJA 1: PODSUMOWANIE ===
    elements.append(Paragraph('PODSUMOWANIE PROWIZJI', section_style))
    elements.append(Spacer(1, 8))
    
    # Tabela podsumowania
    data_summary = [['Grupa zajęciowa', 'Suma kwot', '10% z sumy']]
    total_kwot = 0
    total_prowizji = 0
    
    for _, row in df_handlowiec.iterrows():
        grupa_para = Paragraph(str(row['Grupa zajęciowa']), cell_style)
        data_summary.append([
            grupa_para,
            f"{row['Suma kwot']:.2f} PLN",
            f"{row['10% z sumy']:.2f} PLN"
        ])
        total_kwot += row['Suma kwot']
        total_prowizji += row['10% z sumy']
    
    data_summary.append([
        Paragraph('<b>SUMA CAŁKOWITA:</b>', cell_style_bold),
        f"{total_kwot:.2f} PLN",
        f"{total_prowizji:.2f} PLN"
    ])
    
    col_widths_summary = [150*mm, 55*mm, 55*mm]
    table_summary = Table(data_summary, colWidths=col_widths_summary, repeatRows=1)
    table_summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), DEFAULT_FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -2), DEFAULT_FONT),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d9e1f2')),
        ('FONTNAME', (0, -1), (-1, -1), DEFAULT_FONT_BOLD),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1f4788')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1f4788'))
    ]))
    
    elements.append(table_summary)
    elements.append(Spacer(1, 30))
    
    # === SEKCJA 2: SZCZEGÓŁY WPŁAT ===
    elements.append(PageBreak())
    elements.append(Paragraph('SZCZEGÓŁY WPŁAT', section_style))
    elements.append(Spacer(1, 8))
    
    if not df_szczegoly.empty:
        # Wybór kolumn do wyświetlenia
        columns_to_show = ['data płatności', 'Grupa zajęciowa', 'kwota', 'imię i nazwisko ucznia']
        available_columns = [col for col in columns_to_show if col in df_szczegoly.columns]
        
        if available_columns:
            data_details = [available_columns]
            
            for _, row in df_szczegoly.iterrows():
                row_data = []
                for col in available_columns:
                    value = row[col]
                    if col == 'data płatności':
                        value = value.strftime('%d.%m.%Y %H:%M') if pd.notna(value) else ''
                    elif col == 'kwota':
                        value = f"{value:.2f} PLN" if pd.notna(value) else ''
                    else:
                        value = Paragraph(str(value) if pd.notna(value) else '', cell_style)
                    row_data.append(value)
                data_details.append(row_data)
            
            # Dopasowanie szerokości kolumn
            col_widths_details = [45*mm, 90*mm, 40*mm, 80*mm][:len(available_columns)]
            
            table_details = Table(data_details, colWidths=col_widths_details, repeatRows=1)
            table_details.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), DEFAULT_FONT_BOLD),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('FONTNAME', (0, 1), (-1, -1), DEFAULT_FONT),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1f4788')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f5f5'), colors.white])
            ]))
            
            elements.append(table_details)
    else:
        elements.append(Paragraph('Brak szczegółowych danych wpłat.', cell_style))
    
    doc.build(elements)

def create_handlowiec_pdf_from_excel(filepath, df_podsumowanie, df_szczegoly, handlowiec_name):
    """Tworzenie PDF z wgranego pliku Excel handlowca"""
    
    doc = SimpleDocTemplate(
        filepath,
        pagesize=landscape(A4),
        topMargin=15*mm,
        bottomMargin=20*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Logo
    logo_path = os.path.join('static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=40*mm, height=40*mm)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 5))
        except:
            pass
    
    # Style
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontName=DEFAULT_FONT_BOLD,
        fontSize=18, alignment=TA_CENTER, spaceAfter=12, textColor=colors.HexColor('#1f4788')
    )
    subtitle_style = ParagraphStyle(
        'SubTitle', parent=styles['Heading2'], fontName=DEFAULT_FONT_BOLD,
        fontSize=14, alignment=TA_CENTER, spaceAfter=8, textColor=colors.HexColor('#333333')
    )
    section_style = ParagraphStyle(
        'SectionHeader', parent=styles['Heading2'], fontName=DEFAULT_FONT_BOLD,
        fontSize=14, textColor=colors.HexColor('#1f4788'), spaceAfter=10
    )
    cell_style = ParagraphStyle(
        'CellStyle', parent=styles['Normal'], fontName=DEFAULT_FONT,
        fontSize=9, leading=11
    )
    cell_style_bold = ParagraphStyle(
        'CellStyleBold', parent=styles['Normal'], fontName=DEFAULT_FONT_BOLD,
        fontSize=10, textColor=colors.HexColor('#1f4788'), leading=12
    )
    
    # Tytuł
    elements.append(Paragraph('RAPORT WPŁAT (POPRAWIONY)', title_style))
    elements.append(Paragraph(f'Handlowiec: {handlowiec_name}', subtitle_style))
    elements.append(Spacer(1, 20))
    
    # === SEKCJA 1: PODSUMOWANIE ===
    elements.append(Paragraph('PODSUMOWANIE PROWIZJI', section_style))
    elements.append(Spacer(1, 8))
    
    # Usunięcie ostatniego wiersza jeśli zawiera "SUMA CAŁKOWITA"
    df_podsumowanie_clean = df_podsumowanie.copy()
    if not df_podsumowanie_clean.empty:
        last_row_first_col = str(df_podsumowanie_clean.iloc[-1, 0]).strip().upper()
        if 'SUMA' in last_row_first_col or 'CAŁKOWITA' in last_row_first_col:
            df_podsumowanie_clean = df_podsumowanie_clean[:-1]
    
    # Tabela podsumowania
    data_summary = [df_podsumowanie_clean.columns.tolist()]
    total_kwot = 0
    total_prowizji = 0
    
    for _, row in df_podsumowanie_clean.iterrows():
        row_data = []
        for idx, col in enumerate(df_podsumowanie_clean.columns):
            value = row[col]
            if isinstance(value, (int, float)):
                row_data.append(f"{value:.2f} PLN")
                if idx == 1:
                    total_kwot += value
                elif idx == 2:
                    total_prowizji += value
            else:
                row_data.append(Paragraph(str(value) if pd.notna(value) else '', cell_style))
        data_summary.append(row_data)
    
    # Dodanie sumy całkowitej (obliczonej, nie z Excela)
    data_summary.append([
        Paragraph('<b>SUMA CAŁKOWITA:</b>', cell_style_bold),
        f"{total_kwot:.2f} PLN",
        f"{total_prowizji:.2f} PLN"
    ])
    
    col_widths_summary = [150*mm, 55*mm, 55*mm]
    table_summary = Table(data_summary, colWidths=col_widths_summary, repeatRows=1)
    table_summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), DEFAULT_FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -2), DEFAULT_FONT),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d9e1f2')),
        ('FONTNAME', (0, -1), (-1, -1), DEFAULT_FONT_BOLD),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1f4788')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1f4788'))
    ]))
    
    elements.append(table_summary)
    elements.append(Spacer(1, 30))
    
    # === SEKCJA 2: SZCZEGÓŁY WPŁAT ===
    elements.append(PageBreak())
    elements.append(Paragraph('SZCZEGÓŁY WPŁAT', section_style))
    elements.append(Spacer(1, 8))
    
    if not df_szczegoly.empty:
        # Wybór kolumn do wyświetlenia
        columns_to_show = ['data płatności', 'Grupa zajęciowa', 'kwota', 'imię i nazwisko ucznia']
        available_columns = [col for col in columns_to_show if col in df_szczegoly.columns]
        
        if available_columns:
            data_details = [available_columns]
            
            for _, row in df_szczegoly.iterrows():
                row_data = []
                for col in available_columns:
                    value = row[col]
                    if col == 'data płatności':
                        value = value.strftime('%d.%m.%Y %H:%M') if pd.notna(value) else ''
                    elif col == 'kwota':
                        value = f"{value:.2f} PLN" if pd.notna(value) else ''
                    else:
                        value = Paragraph(str(value) if pd.notna(value) else '', cell_style)
                    row_data.append(value)
                data_details.append(row_data)
            
            # Dopasowanie szerokości kolumn
            col_widths_details = [45*mm, 85*mm, 40*mm, 85*mm][:len(available_columns)]
            
            table_details = Table(data_details, colWidths=col_widths_details, repeatRows=1)
            table_details.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), DEFAULT_FONT_BOLD),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('FONTNAME', (0, 1), (-1, -1), DEFAULT_FONT),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1f4788')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f5f5'), colors.white])
            ]))
            
            elements.append(table_details)
    else:
        elements.append(Paragraph('Brak szczegółowych danych wpłat.', cell_style))
    
    doc.build(elements)

def format_excel(filepath, df_grouped, df_handlowcy_sum):
    """Formatowanie pliku Excel z automatycznym dopasowaniem szerokości kolumn i sumami"""
    wb = openpyxl.load_workbook(filepath)
    
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    sum_font = Font(bold=True, size=11, color="1F4788")
    sum_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = border
                if cell.column in [2, 3]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00 "PLN"'
        
        for column_cells in ws.columns:
            length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        if len(cell_value) > length:
                            length = len(cell_value)
                except:
                    pass
            adjusted_width = min(length + 4, 80)
            ws.column_dimensions[column].width = adjusted_width
        
        ws.freeze_panes = 'A2'
    
    # Sumy w PODSUMOWANIE GRUP
    ws_groups = wb['PODSUMOWANIE GRUP']
    last_row = ws_groups.max_row + 1
    
    suma_kwot = df_grouped['Suma kwot'].sum()
    suma_prowizji_groups = df_grouped['10% z sumy'].sum()
    
    ws_groups[f'A{last_row}'] = 'SUMA CAŁKOWITA:'
    ws_groups[f'B{last_row}'] = suma_kwot
    ws_groups[f'C{last_row}'] = suma_prowizji_groups
    
    for col in ['A', 'B', 'C', 'D']:
        cell = ws_groups[f'{col}{last_row}']
        cell.font = sum_font
        cell.fill = sum_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col in ['B', 'C']:
            cell.number_format = '#,##0.00 "PLN"'
    
    # Sumy w PODSUMOWANIE HANDLOWCÓW
    ws_handlowcy = wb['PODSUMOWANIE HANDLOWCÓW']
    last_row_h = ws_handlowcy.max_row + 1
    
    suma_prowizji_handlowcy = df_handlowcy_sum['Suma prowizji (10%)'].sum()
    
    ws_handlowcy[f'A{last_row_h}'] = 'SUMA CAŁKOWITA:'
    ws_handlowcy[f'B{last_row_h}'] = suma_prowizji_handlowcy
    
    for col in ['A', 'B']:
        cell = ws_handlowcy[f'{col}{last_row_h}']
        cell.font = sum_font
        cell.fill = sum_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col == 'B':
            cell.number_format = '#,##0.00 "PLN"'
    
    wb.save(filepath)

def create_pdf(output_path, df_grouped, df_handlowcy_sum, date_from, date_to, time_from, time_to):
    """Tworzenie głównego raportu PDF"""
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        topMargin=15*mm,
        bottomMargin=20*mm,
        leftMargin=15*mm,
        rightMargin=15*mm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    logo_path = os.path.join('static', 'logo.png')
    if os.path.exists(logo_path):
        try:
            logo = RLImage(logo_path, width=40*mm, height=40*mm)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 5))
        except:
            pass
    
    title_style = ParagraphStyle(
        'CustomTitle', parent=styles['Heading1'], fontName=DEFAULT_FONT_BOLD,
        fontSize=18, alignment=TA_CENTER, spaceAfter=12, textColor=colors.HexColor('#1f4788')
    )
    date_style = ParagraphStyle(
        'DateStyle', parent=styles['Normal'], fontName=DEFAULT_FONT,
        fontSize=11, alignment=TA_CENTER, spaceAfter=20
    )
    section_style = ParagraphStyle(
        'SectionHeader', parent=styles['Heading2'], fontName=DEFAULT_FONT_BOLD,
        fontSize=14, textColor=colors.HexColor('#1f4788'), spaceAfter=10
    )
    cell_style = ParagraphStyle(
        'CellStyle', parent=styles['Normal'], fontName=DEFAULT_FONT,
        fontSize=9, leading=11
    )
    cell_style_bold = ParagraphStyle(
        'CellStyleBold', parent=styles['Normal'], fontName=DEFAULT_FONT_BOLD,
        fontSize=10, textColor=colors.HexColor('#1f4788'), leading=12
    )
    
    elements.append(Paragraph('RAPORT WPŁAT', title_style))
    elements.append(Spacer(1, 10))
    
    date_range = f"Okres: {date_from} {time_from} - {date_to} {time_to}"
    elements.append(Paragraph(date_range, date_style))
    elements.append(Spacer(1, 15))
    
    # TABELA PODSUMOWANIA GRUP
    elements.append(Paragraph('PODSUMOWANIE GRUP', section_style))
    elements.append(Spacer(1, 8))
    
    data_groups = [['Grupa zajęciowa', 'Suma kwot', '10% z sumy', 'Handlowiec']]
    total_kwot = 0
    total_prowizji = 0
    
    for _, row in df_grouped.iterrows():
        grupa_para = Paragraph(str(row['Grupa zajęciowa']), cell_style)
        handlowiec_para = Paragraph(str(row['Handlowiec']), cell_style)
        data_groups.append([
            grupa_para,
            f"{row['Suma kwot']:.2f} PLN",
            f"{row['10% z sumy']:.2f} PLN",
            handlowiec_para
        ])
        total_kwot += row['Suma kwot']
        total_prowizji += row['10% z sumy']
    
    data_groups.append([
        Paragraph('<b>SUMA CAŁKOWITA:</b>', cell_style_bold),
        f"{total_kwot:.2f} PLN",
        f"{total_prowizji:.2f} PLN",
        ''
    ])
    
    col_widths = [110*mm, 45*mm, 45*mm, 65*mm]
    table_groups = Table(data_groups, colWidths=col_widths, repeatRows=1)
    table_groups.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (2, 0), 'RIGHT'),
        ('ALIGN', (3, 0), (3, 0), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), DEFAULT_FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (2, -1), 'RIGHT'),
        ('ALIGN', (3, 1), (3, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -2), DEFAULT_FONT),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d9e1f2')),
        ('FONTNAME', (0, -1), (-1, -1), DEFAULT_FONT_BOLD),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1f4788')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1f4788'))
    ]))
    
    elements.append(table_groups)
    elements.append(Spacer(1, 25))
    
    # TABELA PODSUMOWANIA HANDLOWCÓW
    elements.append(Paragraph('PODSUMOWANIE HANDLOWCÓW', section_style))
    elements.append(Spacer(1, 8))
    
    data_handlowcy = [['Handlowiec', 'Suma prowizji (10%)']]
    total_handlowcy = 0
    
    for _, row in df_handlowcy_sum.iterrows():
        handlowiec_para = Paragraph(str(row['Handlowiec']), cell_style)
        data_handlowcy.append([
            handlowiec_para,
            f"{row['Suma prowizji (10%)']:.2f} PLN"
        ])
        total_handlowcy += row['Suma prowizji (10%)']
    
    data_handlowcy.append([
        Paragraph('<b>SUMA CAŁKOWITA:</b>', cell_style_bold),
        f"{total_handlowcy:.2f} PLN"
    ])
    
    table_handlowcy = Table(data_handlowcy, colWidths=[150*mm, 80*mm], repeatRows=1)
    table_handlowcy.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), DEFAULT_FONT_BOLD),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 1), (-1, -2), DEFAULT_FONT),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#d9e1f2')),
        ('FONTNAME', (0, -1), (-1, -1), DEFAULT_FONT_BOLD),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.HexColor('#1f4788')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1.5, colors.HexColor('#1f4788'))
    ]))
    
    elements.append(table_handlowcy)
    doc.build(elements)

@app.route('/download/<filename>')
def download_file(filename):
    filepath = os.path.join('static', filename)
    return send_file(filepath, as_attachment=True)

@app.route('/manage_handlowcy')
def manage_handlowcy():
    return render_template('handlowcy.html')

@app.route('/get_handlowcy')
def get_handlowcy():
    try:
        df = pd.read_excel(HANDLOWCY_FILE)
        return jsonify(df.to_dict('records'))
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update_handlowcy', methods=['POST'])
def update_handlowcy():
    try:
        data = request.json
        df = pd.DataFrame(data)
        df.to_excel(HANDLOWCY_FILE, index=False)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5100, debug=True)
